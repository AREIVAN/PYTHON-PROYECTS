import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re

import os

nombre_archivo = 'datos.xlsx'
if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active

# crear libro excel
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Telefono", "Direccion"])


def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    mail = entry_mail.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    if not nombre or not edad or not mail or not telefono or not direccion:
        messagebox.showwarning(
            "Campos vacios", "Por favor llene todos los campos")
        return

    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(
            "Error", "Edad y telefono deben ser numeros papu")
        return

    try:
        if "@" not in mail or "." not in mail:
            messagebox.showwarning("Error", "Correo invalido")
            return

        ws.append([nombre, edad, mail, telefono, direccion])
        wb.save(nombre_archivo)
        messagebox.showinfo("Guardado", "Datos guardados correctamente")

        entry_nombre.delete(0, tk.END)
        entry_edad.delete(0, tk.END)
        entry_mail.delete(0, tk.END)
        entry_telefono.delete(0, tk.END)
        entry_direccion.delete(0, tk.END)
    finally:
        pass


root = tk.Tk()
root.title("Formulario de entrada de datos")
root.configure(bg='#4B6587')  # fondo de la app
label_style = {"bg": '#4B6587', "fg": "white"}  # fondo de las letras
entry_style = {"bg": '#D3D3D3', "fg": "black"}  # fondo de la casilla

# letras de nombre
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
# insertar nombre
# primero la entrada * estilo creado antes
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)  # grid y es donde ira


label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)

entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)


label_mail = tk.Label(root, text="Mail", **label_style)
label_mail.grid(row=2, column=0, padx=10, pady=5)

entry_mail = tk.Entry(root, **entry_style)
entry_mail.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)

entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Direccion", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)

entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

boton_guardar = tk.Button(root, text="Guardar",
                          command=guardar_datos, bg='#6D8299', fg='white')
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)
root.mainloop()
